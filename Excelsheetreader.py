# import module
import openpyxl 
# iterating the excel sheet
def datarecord():
    # import dataset from user
    file=input("Name of file: ")
    dtset = openpyxl.load_workbook(file)
    record = dtset.active
    datarecd=[]
    for i in range(1, record.max_row+1):
        print("\n")
        for j in range(1, record.max_column+1):
            col_obj = record.cell(row=i, column=j)
            print(col_obj.value, end=" ")
            datarecd.append(col_obj)
    return datarecd
datarecord()